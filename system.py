from tkinter import *
from tkinter import ttk
import datetime
import openpyxl
from openpyxl import Workbook


tk = Tk()
tk.title('electric machins shop')
tk.geometry('683x911')
tk.resizable(False,False)

#=========== الإطار الأول ===========

Fr1 = Frame(tk,bg='silver',width=391,height=875)
Fr1.place(x=1,y=1)

title1 = Label(Fr1,text = '===== shop products =====',font = ('Tajawal',13),fg='white',bg='#191d51').place(x=31,y=1)

#===== إطار المنتجات =====




menul = {
     0:['لابتوب',6555],
     1:['ميكرويييف',1511],
     2:['غسالة',3109],
     3:['إناء طهي كهربائي',500],
     4:['ثلاجة',7255],
     5:['تلفزيون',9590],
     6:['هاتف',1100],
     7:['سماعات',855],
     8:['كابل توصيل',355],
     9:['مكيف',7400],
     10:['مروحة',1566],
     11:['آلة قهوة',777]

}

time = datetime.datetime.now()
date = time.strftime('%Y-%m-%d')


wr = Workbook()
eng = wr.active

eng.title='Customer'
eng['A1']='Name'
eng['B1']='Mobile Number'
eng['C1']='Card’s Number'
eng['D1']='Adress'
eng['E1']='Products Price'
eng['F1']='Date Salling'

wr.save('shop.xlsx')



def close():
   tk.destroy()



   
def billing():
   
   global entry_n
   global entry_num
   global entry_cn
   global entry_ad
   global entry_pr
   global entry_date
   
   tk.geometry('683x1311')
   
#========== إطار بيانات المشتري ===========
   
   Fr3 = Frame(tk,bg='#5f7161',bd=2,relief=GROOVE,width=699,height=456)
   Fr3.place(x=1,y=895)
   
   list =[]
   
   labe_t = Label(Fr3,text='===== يرتشملا تانايب =====',width=25,height=1,font=('Tajawal',14),bd=1,relief=SOLID,fg='white',bg='#191d51')
   labe_t.place(x=177,y=3)
   
   labe_n = Label(Fr3,text=':  مدختسملا مسا  ',width=15,height=1,font=('Tajawal',14),fg='black',bg='#5f7161')
   labe_n.place(x=471,y=77)
   
   entry_n = Entry(Fr3,width=21,font=('Tajawal',14),bd=1,relief=SOLID,fg='black',justify=CENTER)
   entry_n.place(x=185,y=77)
   
   
   labe_num = Label(Fr3,text='  :   مدختسملا مقر   ',width=15,height=1,font=('Tajawal',14),fg='black',bg='#5f7161')
   labe_num.place(x=471,y=127)
   
   entry_num = Entry(Fr3,width=21,font=('Tajawal',14),bd=1,relief=SOLID,fg='black',justify=CENTER)
   entry_num.place(x=185,y=127)
   
   
   labe_cn = Label(Fr3,text=':   ةقاطبلا مقر    ',width=15,height=1,font=('Tajawal',14),fg='black',bg='#5f7161')
   labe_cn.place(x=471,y=177)
   
   entry_cn = Entry(Fr3,width=21,font=('Tajawal',14),bd=1,relief=SOLID,fg='black',justify=CENTER)
   entry_cn.place(x=185,y=177)
   
   
   labe_ad = Label(Fr3,text=':يرتشملا ناونع',width=15,height=1,font=('Tajawal',14),fg='black',bg='#5f7161')
   labe_ad.place(x=471,y=227)
   
   entry_ad = Entry(Fr3,width=21,font=('Tajawal',14),bd=1,relief=SOLID,fg='black',justify=CENTER)
   entry_ad.place(x=185,y=227)
   
   labe_pr = Label(Fr3,text=':تايرتشملا رعس',width=15,height=1,font=('Tajawal',14),fg='black',bg='#5f7161')
   labe_pr.place(x=471,y=277)
   
   entry_pr = Entry(Fr3,width=21,font=('Tajawal',14),bd=1,relief=SOLID,fg='black',justify=CENTER)
   entry_pr.place(x=185,y=277)
   
   labe_date = Label(Fr3,text=':ءارشلا خيرات',width=15,height=1,font=('Tajawal',14),fg='black',bg='#5f7161')
   labe_date.place(x=471,y=327)
   
   entry_date = Entry(Fr3,width=21,font=('Tajawal',14),bd=1,relief=SOLID,fg='black', justify=CENTER)
   entry_date.place(x=185,y=327)
   
   
   
   save_button =Button(Fr3,width=17,text='ةروـتافلا ظـفـح',cursor='hand2',bd=1,relief=SOLID,bg='#EDDBC0',command=file)
   save_button.place(x=25,y=277)
   
   del_button =Button(Fr3,width=17,text='لوـقـحـلا غارـفا',cursor='hand2',bd=1,relief=SOLID,bg='#EDDBC0',command=clr)
   del_button.place(x=25,y=317)
   
  # searchu_button =Button(Fr3,width=17,text='مدـخـتـسـم نـع ثـحـب',cursor='hand2',bd=1,relief=SOLID,bg='#EDDBC0',command=ser_u)
  # searchu_button.place(x=25,y=357)
   
   #delf_button =Button(Fr3,width=17,text='ةروـتاـف فـذـح',cursor='hand2',bd=1,relief=SOLID,bg='#EDDBC0')
  # delf_button.place(x=25,y=397)
   
   total_p=0
   for itm in tv.get_children():
      tv.delete(itm)
   for i in range(len(snx)):
      if (int(snx[i].get()) > 0):
         price = int(snx[i].get()) * menul[i][1]
         total_p = total_p + price
         val_int = (str(menul[i][1]) , str(snx[i].get()),str(price))
         tv.insert('','end',iid=i,text=menul[i][0],values=val_int)
   
   to = total_p
   entry_pr.insert('1',str(total_p)+'$')
   entry_date.insert('1',str(date))
   
   
def clr():
  for de in tv.get_children():
     tv.delete(de)
     
  entry_n.delete('0',END)
  entry_num.delete('0',END)
  entry_cn.delete('0',END)
  entry_ad.delete('0',END)
  entry_pr.delete('0',END)
  entry_date.delete('0',END)
  
def delete():
   for d in tv.get_children():
      tv.delete(d)
   
   entry_n.delete('0',END)
   entry_num.delete('0',END)
   entry_cn.delete('0',END)
   entry_ad.delete('0',END)
   entry_pr.delete('0',END)
   entry_date.delete('0',END)
        
#متغير حفظ الفاتورة  

def file():
   global xls
   
   entry1 = entry_n.get()
   entry2 = entry_num.get()
   entry3 = entry_cn.get()
   entry4 = entry_ad.get()
   entry5 = entry_pr.get()
   entry6 = entry_date.get()
   
   
   xls = openpyxl.load_workbook('shop.xlsx')
   act = xls.active
   
   act.cell(column=1,row = act.max_row +1,value=entry1)
   act.cell(column=1,row = act.max_row ,value=entry2)
   act.cell(column=1,row = act.max_row ,value=entry3)
   act.cell(column=1,row = act.max_row ,value=entry4)
   act.cell(column=1,row = act.max_row ,value=entry5)
   act.cell(column=1,row = act.max_row ,value=entry6)
   xls.save('shop.xlsx')



#أسماء المنتجات

pr1 = Button(Fr1,width=7,height=4,bg='#EFEAD8',font =('Tajawal',13),bd=1,relief=SOLID,cursor='hand2',text='لابتوب',compound=TOP)
pr1.place(x=16,y=51)
pr2 = Button(Fr1,bg='#EFEAD8',font =('Tajawal',13),bd=1,relief=SOLID,cursor='hand2',text='ميكرووييف',compound=TOP,width=7,height=4)
pr2.place(x=141,y=51)
pr3 = Button(Fr1,bg='#EFEAD8',font =('Tajawal',13),bd=1,relief=SOLID,cursor='hand2',text='غسالة',compound=TOP,width=7,height=4)
pr3.place(x=266,y=51)

pr4 = Button(Fr1,bg='#EFEAD8',font =('Tajawal',13),bd=1,relief=SOLID,cursor='hand2',text='إناء كهربائي',compound=TOP,width=7,height=4)
pr4.place(x=16,y=241)
pr5 = Button(Fr1,bg='#EFEAD8',font =('Tajawal',13),bd=1,relief=SOLID,cursor='hand2',text='ثلاجة',compound=TOP,width=7,height=4)
pr5.place(x=141,y=241)
pr6 = Button(Fr1,bg='#EFEAD8',font =('Tajawal',13),bd=1,relief=SOLID,cursor='hand2',text='تلغزبون',compound=TOP,width=7,height=4)
pr6.place(x=266,y=241)

pr7 = Button(Fr1,bg='#EFEAD8',font =('Tajawal',13),bd=1,relief=SOLID,cursor='hand2',text='هاتف',compound=TOP,width=7,height=4)
pr7.place(x=16,y=432)
pr8 = Button(Fr1,bg='#EFEAD8',font =('Tajawal',13),bd=1,relief=SOLID,cursor='hand2',text='سماعات',compound=TOP,width=7,height=4)
pr8.place(x=141,y=432)
pr9 = Button(Fr1,bg='#EFEAD8',font =('Tajawal',13),bd=1,relief=SOLID,cursor='hand2',text='كابل توصيل',compound=TOP,width=7,height=4)
pr9.place(x=266,y=432)

pr10 = Button(Fr1,bg='#EFEAD8',font =('Tajawal',13),bd=1,relief=SOLID,cursor='hand2',text='مكيف',compound=TOP,width=7,height=4)
pr10.place(x=16,y=622)
pr11 = Button(Fr1,width=7,bg='#EFEAD8',font =('Tajawal',13),bd=1,relief=SOLID,cursor='hand2',text='مروحة',height=4,compound=TOP)
pr11.place(x=141,y=622)
pr12 = Button(Fr1,bg='#EFEAD8',font =('Tajawal',13),bd=1,relief=SOLID,cursor='hand2',text='آلة قهوة',compound=TOP,width=7,height=4)
pr12.place(x=266,y=622)


#===== أزرار القائمة =====

button_p = Button(Fr1,bg='#5f7141',text='ءارش',font=('Tajawal',15),fg='white',bd=1,relief=SOLID,cursor='hand2',compound=TOP,width=3,height=1,command= billing)
button_p.place(x=16,y=821)

button_e = Button(Fr1,bg='#5f7141',text='راجأتسا',font=('Tajawal',15),fg='white',bd=1,relief=SOLID,cursor='hand2',compound=TOP,width=5,height=1,command = billing)
button_e.place(x=92,y=821)

button_n = Button(Fr1,bg='#5f7141',text='چ.ةروتاف',font=('Tajawal',15),fg='white',bd=1,relief=SOLID,cursor='hand2',compound=TOP,width=5,height=1,command=delete)
button_n.place(x=196,y=821)

button_c = Button(Fr1,bg='#5f7141',text='قالغا',font=('Tajawal',15),fg='white',bd=1,relief=SOLID,cursor='hand2',compound=TOP,width=4,height=1,command=close)
button_c.place(x=300,y=821)


#===== حقول الأعداد =====

snx= []

int1 = IntVar()
int2 = IntVar()
int3 = IntVar()
int4 = IntVar()
int5 = IntVar()
int6 = IntVar()
int7 = IntVar()
int8 = IntVar()
int9 = IntVar()
int10 = IntVar()
int11 = IntVar()
int12 = IntVar()



#حقول عدد المشتريات

sp1= Spinbox(Fr1,from_=0,to_=5,font=('Times',14,'normal'),width=7,textvariable=int1)
sp1.place(x=16,y=174)
snx.append(sp1)

sp2= Spinbox(Fr1,from_=0,to_=5,font=('Times',14,'normal'),width=7,textvariable=int2)
sp2.place(x=141,y=174)
snx.append(sp2)

sp3= Spinbox(Fr1,from_=0,to_=5,font=('Times',14,'normal'),width=7,textvariable=int3)
sp3.place(x=266,y=174)
snx.append(sp3)

sp4= Spinbox(Fr1,from_=0,to_=5,font=('Times',14,'normal'),width=7,textvariable=int4)
sp4.place(x=16,y=364)
snx.append(sp4)

sp5= Spinbox(Fr1,from_=0,to_=5,font=('Times',14,'normal'),width=7,textvariable=int5)
sp5.place(x=141,y=364)
snx.append(sp5)

sp6= Spinbox(Fr1,from_=0,to_=5,font=('Times',14,'normal'),width=7,textvariable=int6)
sp6.place(x=266,y=364)
snx.append(sp6)

sp7= Spinbox(Fr1,from_=0,to_=5,font=('Times',14,'normal'),width=7,textvariable=int7)
sp7.place(x=16,y=555)
snx.append(sp7)

sp8= Spinbox(Fr1,from_=0,to_=5,font=('Times',14,'normal'),width=7,textvariable=int8)
sp8.place(x=141,y=555)
snx.append(sp8)

sp9= Spinbox(Fr1,from_=0,to_=5,font=('Times',14,'normal'),width=7,textvariable=int9)
sp9.place(x=266,y=555)
snx.append(sp9)

sp10= Spinbox(Fr1,from_=0,to_=5,font=('Times',14,'normal'),width=7,textvariable=int10)
sp10.place(x=16,y=745)
snx.append(sp10)

sp11= Spinbox(Fr1,from_=0,to_=5,font=('Times',14,'normal'),width=7,textvariable=int11)
sp11.place(x=141,y=745)
snx.append(sp11)

sp12= Spinbox(Fr1,from_=0,to_=5,font=('Times',14,'normal'),width=7,textvariable=int12)
sp12.place(x=266,y=745)
snx.append(sp12)



#=========== إطار الفاتورة ===========

Fr2 = Frame(tk,bg='light gray',width=303,height=875)
Fr2.place(x=401,y=1)

tv = ttk.Treeview(Fr2,selectmode ='browse')
tv.place(x=1,y=1,width=307,height=875)

tv['columns'] = ('1','2','3')

tv.column('#0',width=81,anchor='c')
tv.column('1',width=51,anchor='c')
tv.column('2',width=41,anchor='c')
tv.column('3',width=71,anchor='c')

tv.heading('#0',text='داوملا',anchor='c')
tv.heading('1',text='رعسلا',anchor='c')
tv.heading('2',text='ددعلا',anchor='c')
tv.heading('3',text='يلكلا باسحلا',anchor='c')



tk.mainloop()
